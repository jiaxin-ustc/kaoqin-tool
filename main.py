import pandas as pd
from datetime import time, datetime
import gradio as gr
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

import sys
import os

# 修复打包后的环境问题
if getattr(sys, 'frozen', False):
    # 如果是打包后的 exe
    if sys.stdout is None:
        sys.stdout = open(os.devnull, 'w')
    if sys.stderr is None:
        sys.stderr = open(os.devnull, 'w')

# 考勤规则配置
# 上午打卡时间：8:00之前、12:00之后；
# 下午打卡时间：13:00之前、15:30之后
class AttendanceConfig:
    MORNING_START = time(8, 0)      # 上午上班时间
    MORNING_END = time(12, 0)       # 上午下班时间
    AFTERNOON_START = time(13, 00)  # 下午上班时间
    AFTERNOON_END = time(15, 30)    # 下午下班时间
    ALLOWED_LATE_MINUTES = 0       # 允许迟到分钟数
    ALLOWED_EARLY_MINUTES = 0      # 允许早退分钟数

def time_add_minutes(t, minutes):
    """时间加减分钟"""
    if not isinstance(t, time):
        return None
    total = t.hour * 60 + t.minute + minutes
    return time(total // 60 % 24, total % 60)

def process_attendance_data(df):
    """处理考勤数据主逻辑"""
    # 提取日期行（假设第3行是日期行）
    date_row = df.columns.values[2:] #date_row = df.iloc[0, 2:].values

    results = []
    for row_idx in range(0, len(df)):
        row = df.iloc[row_idx]
        employee = row[0]
        department = row[1]
        
        for col_idx in range(2, min(len(row), len(date_row)+2)):
            date_str = date_row[col_idx-2]
            print("date_str:",date_str)
            date = date_str
            # date = f"2025-06-{date_str}" if str(date_str).isdigit() else f"特殊日期_{col_idx}"
            
            # 处理时间数据
            times = process_cell_times(row[col_idx])
            periods = assign_time_periods(times)
            status = check_attendance_status(periods)
            
            record = {
                'employee': employee,
                'department': department,
                'date': date,
                **periods,
                **status
            }
            results.append(record)

    return pd.DataFrame(results)

def parse_time(time_str):
    """将时间字符串转换为time对象"""
    if pd.isna(time_str) or not isinstance(time_str, str):
        return None
    
    try:
        # 清理字符串并提取时间部分
        time_str = re.sub(r'[^\d:]', '', time_str.strip())
        if ':' in time_str:
            parts = time_str.split(':')
            if len(parts) >= 2:
                return time(int(parts[0]), int(parts[1]))
        elif len(time_str) == 4 and time_str.isdigit():
            return time(int(time_str[:2]), int(time_str[2:]))
    except:
        return None
    return None

def process_cell_times(cell_value):
    """处理单元格内的多个时间记录"""
    if pd.isna(cell_value) or cell_value == '':
        return []
    
    # 统一处理换行符和空格
    if isinstance(cell_value, str):
        times = []
        # 先尝试按换行符分割
        for t in cell_value.split('\n'):
            t = t.strip()
            if t:
                # 再尝试按空格分割（处理"07:35 12:08"格式）
                for subt in t.split():
                    parsed = parse_time(subt)
                    if parsed:
                        times.append(parsed)
        return times
    return []

def assign_time_periods(times):
    """
    智能分配时间点到对应时间段（枚举所有情况）
        上午打卡时间：8:00之前、12:00之后；
        下午打卡时间：13:00之前、15:30之后
    """
    config = AttendanceConfig()
    # 过滤无效时间并按时间排序
    valid_times = sorted([t for t in times if isinstance(t, time)])
    print("valid_times:", valid_times)
    num_times = len(valid_times)
    
    # 初始化结果
    result = {
        'morning_start': None,
        'morning_end': None,
        'afternoon_start': None,
        'afternoon_end': None
    }
    
    # 根据时间点数量处理不同情况
    if num_times == 0:
        # 无有效时间点，全部缺卡
        return result
    # MORNING_START 
    elif num_times == 1:
        # 只有一个时间点
        t = valid_times[0]
        if t <= config.MORNING_START:
            result['morning_start'] = t
        elif config.MORNING_END <= t <= config.AFTERNOON_START:
            # 可能在上午下班和下午上班之间
            result['morning_end'] = t
        elif config.AFTERNOON_START < t < config.AFTERNOON_END:
            result['afternoon_start'] = t
        elif t >= config.AFTERNOON_END:
            result['afternoon_end'] = t
        else:
            # 无法确定时间段，按时间顺序分配
            result['morning_start'] = t
    
    elif num_times == 2:
        t1, t2 = valid_times
        print("num_times == 2:",t1, t2)
        print(config.MORNING_START, config.MORNING_END, config.AFTERNOON_START, config.AFTERNOON_END)
        # 可能的情况：上午两次打卡、下午两次打卡、跨午休打卡
        if t1 <= config.MORNING_START and t2 >= config.MORNING_END: ##TODO: 这里需要修改

            # 上午两次打卡
            result['morning_start'] = t1
            result['morning_end'] = t2
        elif t1 <= config.AFTERNOON_START and t2 >= config.AFTERNOON_END:
            # 下午两次打卡
            result['afternoon_start'] = t1
            result['afternoon_end'] = t2
        elif t1 < config.MORNING_END and t2 > config.AFTERNOON_START:
            # 跨午休打卡
            result['morning_end'] = t1
            result['afternoon_start'] = t2
        else:
            # 默认分配
            result['morning_start'] = t1
            result['morning_end'] = t2
    
    elif num_times == 3:
        t1, t2, t3 = valid_times
        # 可能缺上午上班、上午下班、下午上班或下午下班
        if t1 <= config.MORNING_START and t2 >= config.MORNING_END and t3 >= config.AFTERNOON_END:
            # 缺下午上班
            result['morning_start'] = t1
            result['morning_end'] = t2
            result['afternoon_end'] = t3
        elif t1 <= config.MORNING_START and t2 <= config.AFTERNOON_START and t3 >= config.AFTERNOON_END:
            # 缺上午下班
            result['morning_start'] = t1
            result['afternoon_start'] = t2
            result['afternoon_end'] = t3
        elif t1 >= config.MORNING_END and t2 <= config.AFTERNOON_START and t3 >= config.AFTERNOON_END:
            # 缺上午上班
            result['morning_end'] = t1
            result['afternoon_start'] = t2
            result['afternoon_end'] = t3
        elif t1 <= config.MORNING_START and t2 >= config.MORNING_END and t3 <= config.AFTERNOON_START:
            # 缺下午下班
            result['morning_start'] = t1
            result['morning_end'] = t2
            result['afternoon_start'] = t3
        else:
            # 默认分配
            result['morning_start'] = t1
            result['morning_end'] = t2
            result['afternoon_start'] = t3
    
    elif num_times >= 4:
        # 有4个或更多时间点，取前4个
        t1, t2, t3, t4 = valid_times[:4]
        result['morning_start'] = t1
        result['morning_end'] = t2
        result['afternoon_start'] = t3
        result['afternoon_end'] = t4
    
    return result

def check_attendance_status(record):
    """检查考勤状态（优化版）"""
    config = AttendanceConfig()
    status = {
        'late': False,
        'early_leave': False,
        'missing': False,
        'status_details': []
    }
    
    # 检查上午上班
    if record['morning_start']:
        if record['morning_start'] > time_add_minutes(config.MORNING_START, config.ALLOWED_LATE_MINUTES):
            status['late'] = True
            status['status_details'].append('上午迟到')
    else:
        status['missing'] = True
        status['status_details'].append('上午上班缺卡')
    
    # 检查上午下班
    if record['morning_end']:
        if record['morning_end'] < time_add_minutes(config.MORNING_END, -config.ALLOWED_EARLY_MINUTES):
            status['early_leave'] = True
            status['status_details'].append('上午早退')
    else:
        status['missing'] = True
        status['status_details'].append('上午下班缺卡')
    
    # 检查下午上班
    if record['afternoon_start']:
        if record['afternoon_start'] > time_add_minutes(config.AFTERNOON_START, config.ALLOWED_LATE_MINUTES):
            status['late'] = True
            status['status_details'].append('下午迟到')
    else:
        status['missing'] = True
        status['status_details'].append('下午上班缺卡')
    
    # 检查下午下班
    if record['afternoon_end']:
        if record['afternoon_end'] < time_add_minutes(config.AFTERNOON_END, -config.ALLOWED_EARLY_MINUTES):
            status['early_leave'] = True
            status['status_details'].append('下午早退')
    else:
        status['missing'] = True
        status['status_details'].append('下午下班缺卡')
    
    return status

def write_results_back_to_excel(original_file, result_df, output_file=None):
    """
    将考核评估结果写回原表格，并标红存在考勤问题的记录
    
    Args:
        original_file: 原始Excel文件路径
        result_df: 处理后的考勤结果DataFrame
        output_file: 输出文件路径，如果为None则覆盖原文件
    
    Returns:
        str: 处理结果消息
    """
    try:
        print(f"开始处理文件: {original_file}")
        
        # 读取原始Excel文件
        original_df = pd.read_excel(original_file, header=3)
        original_df = original_df.rename(columns={'Unnamed: 0': '姓名', 'Unnamed: 1': '部门'})
        
        # 创建异常记录映射
        # 格式: {(员工姓名, 日期): [异常详情列表]}
        issue_mapping = {}
        
        # 只处理有异常的记录
        abnormal_records = result_df[result_df['status_details'].apply(len) > 0]
        print(f"发现 {len(abnormal_records)} 条异常记录")
        
        for _, record in abnormal_records.iterrows():
            employee = record['employee']
            date = record['date']
            status_details = record['status_details']
            
            issue_mapping[(employee, date)] = status_details
        
        # 使用openpyxl加载工作簿
        wb = load_workbook(original_file)
        ws = wb.active
        
        # 定义红色填充样式
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        red_font = Font(color='FFFFFF')  # 白色字体，便于在红色背景上阅读
        
        # 获取日期行（第4行，索引为3）
        date_row = 4
        date_cols = {}
        
        # 获取日期列映射
        for col in range(3, ws.max_column + 1):  # 从第3列开始（C列）
            cell_value = ws.cell(row=date_row, column=col).value
            if cell_value:
                date_cols[str(cell_value)] = col
        
        print(f"找到 {len(date_cols)} 个日期列")
        
        # 获取员工行映射
        employee_rows = {}
        for row in range(5, ws.max_row + 1):  # 从第5行开始（数据行）
            employee_name = ws.cell(row=row, column=1).value  # 第1列是姓名
            if employee_name:
                employee_rows[str(employee_name)] = row
        
        print(f"找到 {len(employee_rows)} 个员工")
        
        # 标记异常记录
        marked_cells = 0
        for (employee, date), issues in issue_mapping.items():
            if employee in employee_rows and date in date_cols:
                row_idx = employee_rows[employee]
                col_idx = date_cols[date]
                
                # 获取原始单元格
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # 添加异常标记到单元格内容
                original_value = cell.value if cell.value else ""
                issue_text = " | ".join(issues)
                
                # 如果单元格不为空，添加换行符分隔
                if original_value:
                    new_value = f"{original_value}\n[异常: {issue_text}]"
                else:
                    new_value = f"[异常: {issue_text}]"
                
                cell.value = new_value
                
                # 应用红色样式
                cell.fill = red_fill
                cell.font = red_font
                
                marked_cells += 1
                print(f"标记异常: {employee} - {date} - {issue_text}")
            else:
                print(f"未找到对应位置: {employee} - {date}")
        
        # 保存文件
        if output_file is None:
            output_file = original_file.replace('.xlsx', '_marked.xlsx')
        
        wb.save(output_file)
        
        return f"成功标记 {marked_cells} 个异常记录，结果已保存到: {output_file}"
        
    except Exception as e:
        import traceback
        print(f"错误详情: {traceback.format_exc()}")
        return f"写回原表格失败: {str(e)}"

def analyze_attendance(input_file):
    """分析考勤数据主函数"""
    # 读取Excel文件
    try:
        # 跳过前两行标题，保留日期行
        df = pd.read_excel(input_file, header=3) 
        print(df.columns)
        df = df.rename(columns={'Unnamed: 0': '姓名', 'Unnamed: 1': '部门'})
    except Exception as e:
        return f"读取文件失败: {str(e)}", None, None
    
    # 处理考勤数据
    result_df = process_attendance_data(df)
    # result_df.to_excel("result_zhongjian.xlsx", index=False)
    
    if result_df.empty:
        return "未找到有效考勤数据", None, None
    
    # 创建详细报告
    detailed_report = result_df.copy()
    
    # 格式化时间列
    time_cols = ['morning_start', 'morning_end', 'afternoon_start', 'afternoon_end']
    for col in time_cols:
        detailed_report[col] = detailed_report[col].apply(
            lambda x: x.strftime('%H:%M') if isinstance(x, time) else ''
        )
    
    # 重命名列
    detailed_report = detailed_report.rename(columns={
        'employee': '员工姓名',
        'date': '日期',
        'morning_start': '上午上班',
        'morning_end': '上午下班',
        'afternoon_start': '下午上班',
        'afternoon_end': '下午下班',
        'late': '迟到',
        'early_leave': '早退',
        'missing': '缺卡'
    })
    
    # 创建汇总报告
    if '员工姓名' in detailed_report.columns:
        summary = detailed_report.groupby('员工姓名').agg({
            '迟到': 'sum',
            '早退': 'sum',
            '缺卡': 'sum'
        }).reset_index()
        
        # 添加总计行
        total_row = pd.DataFrame({
            '员工姓名': ['总计'],
            '迟到': [summary['迟到'].sum()],
            '早退': [summary['早退'].sum()],
            '缺卡': [summary['缺卡'].sum()]
        })
        summary = pd.concat([summary, total_row], ignore_index=True)
    else:
        summary = pd.DataFrame({"状态": ["无有效员工数据"]})
    
    #只保留detailed_report中,"status_details"非空的行!!!!!!!!!!!!!!
    detailed_report= detailed_report[detailed_report['status_details'].apply(len) > 0]
    # detailed_report.to_excel("kaohe_result.xlsx", index=False)
    
    # 将结果写回原表格并标红异常记录
    write_back_result = write_results_back_to_excel(input_file, result_df)
    
    return "分析完成", detailed_report, summary, write_back_result

# Gradio界面
with gr.Blocks() as demo:
    gr.Markdown("## 考勤异常检测系统")
    with gr.Row():
        file_input = gr.File(label="上传考勤表", file_types=[".xlsx"])
        analyze_btn = gr.Button("分析数据", variant="primary")
    
    output_text = gr.Textbox(label="处理结果")
    write_back_text = gr.Textbox(label="写回原表格结果")
    with gr.Tabs():
        with gr.Tab("详细记录"):
            detail_table = gr.Dataframe(interactive=False)
        with gr.Tab("汇总统计"):
            summary_table = gr.Dataframe(interactive=False)
    
    analyze_btn.click(
        analyze_attendance,
        inputs=file_input,
        outputs=[output_text, detail_table, summary_table, write_back_text]
    )
    #版权信息
    gr.HTML(
        "<div style='text-align:center; color:gray; margin-top:20px;'>版权所有 © 2025 Jiaxin Jiang</div>"
        )

if __name__ == "__main__":
    demo.launch()